import praw
import config
import time
import os
import random
import openpyxl
import datetime
GonderilenPostIdleri = []
Satirsublar = []
Postlinks = []
Hata = "community"

import socket

def Post_gonder(subreddit1,r,username_Used):
        #path = "../POSTLAR.xlsx" 
        wb_obj = openpyxl.load_workbook(path)         
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
        x = datetime.datetime.now()
        x = int(x.strftime("%H"))
#         if(6>=x>=4):
#             time.sleep(10800)    
#         elif(8>=x>=7):   
#             time.sleep(10800)
#         elif(16>=x>=9):   
#             time.sleep(1800) 
        time.sleep(10) # 2 dk post var mı yok mu kontrol eder.
        print ("\n\nPOST VAR MI YOK MU KONTROL EDİYORUM (120 sn)","\n")
        for i in range(1, m_row + 1):
            title=sheet_obj.cell(row = i, column = 1).value
            link=sheet_obj.cell(row = i, column = 2).value
            com_yorumlink=sheet_obj.cell(row = 1, column = 3).value
            com_yorumtg=sheet_obj.cell(row = 1, column = 4).value
            with open("GonderilenPostlar.txt", "r") as file3:  #Paylasilansublar.txt elle oluştur.
                Postlinks = file3.read()
                Postlinks = Postlinks.split("\n")
                Postlinks = list(filter(None, Postlinks))
                if link not in Postlinks :
                    print("\n\n YENİ İÇERİK BULUNDU -> ",link,"\n")
                    with open ("GonderilenPostlar.txt", "a") as f:
                        r.validate_on_submit = True
                        yorum = "#"+com_yorumlink+"\n #Telegram link : "+com_yorumtg
                        submission=r.subreddit(username_Used).submit(title=title, url=link, nsfw=True).reply(body=yorum)                        
                        f.write(link + "\n")
                        time.sleep(5)
                        return
        
def internet(host="8.8.8.8", port=53, timeout=3):
    """
    Host: 8.8.8.8 (google-public-dns-a.google.com)
    OpenPort: 53/tcp
    Service: domain (DNS/TCP)
    """
    try:
        socket.setdefaulttimeout(timeout)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect((host, port))
        return True
    except socket.error as ex:
        print(ex)
        return False

def randomTime():
    randomText1 = random.randint(1,2)
    return  randomText1
def randomLine():
    textim = random.choice(list(open('readme.txt')))
    return textim
def randomLink():
    textlink = random.choice(list(open('justetext.txt')))
    return textlink

def bot_login():
	print ("Logging in...")
	r = praw.Reddit(username = config.username,
				password = config.password,
				client_id = config.client_id,
				client_secret = config.client_secret,
				user_agent = "replay bot by u/e12323wsds34dwqs")
	print ("Logged in!")

	return r



def run_bot(subreddit1, GonderilenPostIdleri,r,username_Used):    
    for submission in subreddit1.hot(limit=3):
        if submission.id not in GonderilenPostIdleri :
            print ("Yeni Post bulundu. Post id : " + submission.id)
     
            with open('subreddits.txt') as file:
                for line in file:
                    print("Subreddit sıraya alındı : " + line.rstrip())
                    with open("Paylasilansublar.txt") as file1:  #Paylasilansublar.txt elle oluştur.
                        Satirsublar = [line.strip() for line in file1]
                    if line.rstrip() not in Satirsublar :
                        submission1 = r.submission(id=submission.id)
#                         submission1.comments.replace_more(limit=1)
#                         comments = submission1.comments.list()
#                         comment_id = comments[0]
#                         comment = r.comment(comment_id)
# 
#                       
#                         yorum = "#[🟥MEG4 L1NK](https://www.reddit.com"+comment_id.permalink+") "
#                         print(yorum)
                        r.validate_on_submit = True
                        #r.subreddit(line.rstrip()).submit(submission.title, url=submission.url, nsfw=True).reply(body=yorum)
                        try:

                            r.subreddit(line.rstrip()).submit(submission.title, url=submission.url) # .reply(body=yorum)
                                            
                        except Exception as e:
                            print(str(e))
                            
                            if ("community" in str(e)  or  "frlair"  in str(e)):
                                with open("subreddits.txt", "r") as f:
                                    lines = f.readlines()
                                with open("subreddits.txt", "w") as f:
                                    for line1 in lines:
                                        if (line1.strip("\n") != line.rstrip()):
                                            f.write(line1)
                                    pass 
                               
                                    
                            else:
                                 try:
                                    r.subreddit(line.rstrip()).submit(submission.title, url=submission.url, nsfw=True) #.reply(body=yorum)
                                 except:
                                    with open("subreddits.txt", "r") as f:
                                        lines = f.readlines()
                                    with open("subreddits.txt", "w") as f:
                                        for line1 in lines:
                                            if (line1.strip("\n") != line.rstrip()):
                                                f.write(line1)
                                        pass
      
                        with open ("Paylasilansublar.txt", "a") as f:
                            f.write(line.rstrip() + "\n")
                        print ("Post Göderimi "+line.rstrip()+" subredditine yapıldı :  " + submission.id+"\n")
                        print ("Crosspost için yeni subreddit bekleniyor (60sn) ", time.sleep(randomTime()))
            
            GonderilenPostIdleri.append(submission.id)
          
            with open ("GonderilenPostIdleri.txt", "a") as f:
                f.write(submission.id + "\n")
                print ("GonderilenPostIdleri.txt'nin içine crosspost yapılan id kaydedildi.")
            print ("Crosspost işlemiş Bitti. " + submission.id)
            with open("Paylasilansublar.txt", "r+") as file2:
                file2.truncate()
                print ("Paylasilansublar.txt'in içi temizlendi.\n")
                print ("Yeni içerik için crosspostu  2 dk beklet. ",time.sleep(10800),"\n")
    print ("Corsspost Yapılan Post idleri : ",GonderilenPostIdleri)
    #Post_gonder(subreddit1,r,username_Used)
    
                
  
def get_saved_comments():

	if not os.path.isfile("GonderilenPostIdleri.txt"):
		GonderilenPostIdleri = []
	else:
		with open("GonderilenPostIdleri.txt", "r") as f:
			GonderilenPostIdleri = f.read()
			GonderilenPostIdleri = GonderilenPostIdleri.split("\n")
			GonderilenPostIdleri = list(filter(None, GonderilenPostIdleri))

	return GonderilenPostIdleri




r = bot_login()
username_Used = "u_"+config.username
subreddit1 = r.subreddit('shitposting') #kendi subreddit sayfanın urlsi gelecek
GonderilenPostIdleri = get_saved_comments()







while True:
    try:
        run_bot(subreddit1, GonderilenPostIdleri,r,username_Used)
    except Exception as e:
        if (internet()):
            run_bot(subreddit1, GonderilenPostIdleri,r,username_Used)
        else:
            print("INTERNET YOK")
            time.sleep(500)
    
        
